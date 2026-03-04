"""
export/excel_exporter.py
===========================
Export hasil analisis Tenrix ke file Excel (.xlsx).

Struktur output:
  Sheet 1: README       — ringkasan sesi, metadata, AI interpretation gabungan
  Sheet 2: Data Profile — stats profil dataset (rows, cols, missing, dll)
  Sheet 3+: per analisis — satu sheet per analisis yang dijalankan

Setiap sheet analisis berisi:
  - Summary table (metrics dari analisis)
  - Data table (hasil detail, kalau ada)
  - AI Interpretation (sebagai teks di bawah tabel)

Dependency: openpyxl (sudah ada di requirements.txt)
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from analysis.statistics import AnalysisResult


# ── Warna tema ────────────────────────────────────────────────────────────────
BLUE_DARK   = "1E3A5F"   # header utama
BLUE_MID    = "2563EB"   # header sekunder
BLUE_LIGHT  = "EFF6FF"   # row highlight
GRAY_LIGHT  = "F8FAFC"   # alternating row
WHITE       = "FFFFFF"
GREEN       = "10B981"
ORANGE      = "D97706"
TEXT_DARK   = "0F172A"
TEXT_GRAY   = "64748B"


# ── Entry point ───────────────────────────────────────────────────────────────

def export_excel(
    results:    list[AnalysisResult],
    profile,
    file_path:  str,
    source_file: str = "",
    guardrails: dict = None,
    exec_summary: str = "",
) -> str:
    """
    Export semua hasil analisis ke .xlsx.
    Return path file yang dibuat.

    profile can be a DataProfile dataclass or a dict (from analysis/profiler.py).
    """
    wb = Workbook()

    # Hapus sheet default
    wb.remove(wb.active)

    # Sheet 1: README
    _write_readme_sheet(wb, results, profile, source_file, exec_summary)

    # Sheet 2: Data Profile
    _write_profile_sheet(wb, profile)
    
    # Sheet 2.5: Data Quality (if guardrails exist)
    if guardrails:
        _write_data_quality_sheet(wb, guardrails)

    # Sheet 3+: satu per analisis
    for i, result in enumerate(results, 1):
        if result.success:
            g = guardrails.get(result.analysis_id) if guardrails else None
            _write_analysis_sheet(wb, result, index=i, guardrail=g)

    # Simpan
    path = Path(file_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))

    return str(path)


# ── Sheet: README ─────────────────────────────────────────────────────────────

def _write_readme_sheet(
    wb:          Workbook,
    results:     list[AnalysisResult],
    profile,
    source_file: str,
    exec_summary: str = "",
) -> None:
    ws = wb.create_sheet("README", 0)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 52

    # Judul
    ws.merge_cells("A1:B1")
    cell = ws["A1"]
    cell.value          = "TENRIX — Data Analysis Report"
    cell.font           = Font(name="Calibri", bold=True, size=16, color=WHITE)
    cell.fill           = PatternFill("solid", fgColor=BLUE_DARK)
    cell.alignment      = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Extract profile values (handle both dict and dataclass)
    row_count = _get_profile_val(profile, "row_count", 0)
    columns = _get_profile_val(profile, "columns", [])
    col_count = len(columns) if isinstance(columns, (list, dict)) else _get_profile_val(profile, "col_count", 0)

    # Metadata
    meta = [
        ("File",       Path(source_file).name if source_file else "—"),
        ("Date",       datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Rows",       f"{row_count:,}"),
        ("Columns",    str(col_count)),
        ("Analyses",   str(sum(1 for r in results if r.success))),
        ("Generated",  "Tenrix CLI"),
    ]
    for r_idx, (key, val) in enumerate(meta, 2):
        k_cell        = ws.cell(row=r_idx, column=1, value=key)
        k_cell.font   = Font(bold=True, color=TEXT_DARK)
        k_cell.fill   = PatternFill("solid", fgColor=BLUE_LIGHT)

        v_cell        = ws.cell(row=r_idx, column=2, value=val)
        v_cell.font   = Font(color=TEXT_DARK)

        ws.row_dimensions[r_idx].height = 18

    # Spacer
    spacer_row = len(meta) + 3
    ws.cell(row=spacer_row, column=1, value="Analyses in this report").font = Font(
        bold=True, size=12, color=BLUE_DARK
    )

    # Daftar analisis
    for i, result in enumerate(results, 1):
        row  = spacer_row + i
        icon = "✅" if result.success else "❌"
        ws.cell(row=row, column=1, value=f"{icon} {result.analysis_name}")
        ws.cell(row=row, column=1).font = Font(color=TEXT_DARK)
        ws.cell(row=row, column=2, value=f"→ Sheet: {_safe_sheet_name(result.analysis_name, i)}")
        ws.cell(row=row, column=2).font = Font(color=TEXT_GRAY)

    _apply_thin_border(ws, 1, 1, len(meta) + 1, 2)
    
    # Executive Summary
    if exec_summary:
        current_row = row + 3
        ws.cell(row=current_row, column=1, value="Executive Summary").font = Font(
            bold=True, size=12, color=BLUE_DARK
        )
        current_row += 1
        
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        exec_cell = ws.cell(row=current_row, column=1)
        exec_cell.value = exec_summary
        exec_cell.alignment = Alignment(wrap_text=True, vertical="top")
        exec_cell.font = Font(color=TEXT_DARK, size=11)
        
        line_count = max(5, len(exec_summary) // 80 + 2)
        ws.row_dimensions[current_row].height = line_count * 15

# ── Sheet: Data Quality (New) ────────────────────────────────────────────────

def _write_data_quality_sheet(wb: Workbook, guardrails: dict) -> None:
    ws = wb.create_sheet("Data Quality")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 60
    
    row = _write_section_header(ws, 1, "Data Quality & Statistical Assumptions")
    row += 1
    
    passed_count = sum(1 for g in guardrails.values() if g.passed)
    total_count = len(guardrails)
    
    ws.cell(row=row, column=1, value="Overall Status").font = Font(bold=True)
    status_text = "✅ GOOD" if passed_count == total_count else f"⚠️ {total_count - passed_count} WARNINGS"
    status_color = GREEN if passed_count == total_count else ORANGE
    ws.cell(row=row, column=2, value=status_text).font = Font(bold=True, color=status_color)
    row += 2
    
    # Headers
    for c, title in enumerate(["Analysis", "Status", "Details/Violations"], 1):
        cell = ws.cell(row=row, column=c, value=title)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE_MID)
    row += 1
    
    for aid, g in guardrails.items():
        ws.cell(row=row, column=1, value=aid)
        
        if g.passed:
            ws.cell(row=row, column=2, value="✅ Passed").font = Font(color=GREEN)
            ws.cell(row=row, column=3, value="All assumptions met")
            row += 1
        else:
            ws.cell(row=row, column=2, value="⚠️ Warning").font = Font(color=ORANGE)
            violations_text = "\n".join([f"- {v.message}" for v in g.violations if not v.passed])
            v_cell = ws.cell(row=row, column=3, value=violations_text)
            v_cell.alignment = Alignment(wrap_text=True)
            ws.row_dimensions[row].height = max(15, 15 * len([v for v in g.violations if not v.passed]))
            row += 1


# ── Sheet: Data Profile ───────────────────────────────────────────────────────

def _write_profile_sheet(wb: Workbook, profile) -> None:
    ws = wb.create_sheet("Data Profile")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    _write_section_header(ws, 1, "Data Profile")

    # Extract values from profile (handle both dict and dataclass)
    row_count = _get_profile_val(profile, "row_count", 0)
    columns = _get_profile_val(profile, "columns", [])
    col_count = len(columns) if isinstance(columns, (list, dict)) else _get_profile_val(profile, "col_count", 0)

    # Get numeric/categorical/date column counts
    numeric_cols = _get_profile_val(profile, "numeric_columns", [])
    categorical_cols = _get_profile_val(profile, "categorical_columns", [])
    date_cols = _get_profile_val(profile, "date_columns", [])

    numeric_count = len(numeric_cols) if isinstance(numeric_cols, list) else numeric_cols
    categorical_count = len(categorical_cols) if isinstance(categorical_cols, list) else categorical_cols
    date_count = len(date_cols) if isinstance(date_cols, list) else date_cols

    missing_count = _get_profile_val(profile, "missing_count",
                                     _get_profile_val(profile, "total_missing", 0))
    duplicate_count = _get_profile_val(profile, "duplicate_count",
                                       _get_profile_val(profile, "duplicate_rows", 0))

    rows = [
        ("Rows",             f"{row_count:,}"),
        ("Columns",          str(col_count)),
        ("Numeric Columns",  str(numeric_count)),
        ("Categorical Cols", str(categorical_count)),
        ("Date Columns",     str(date_count)),
        ("Missing Values",   f"{missing_count:,}"),
        ("Duplicate Rows",   f"{duplicate_count:,}"),
    ]

    for i, (key, val) in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=key).font  = Font(bold=True)
        ws.cell(row=i, column=2, value=val).font  = Font(color=TEXT_DARK)
        if i % 2 == 0:
            for col in [1, 2]:
                ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor=GRAY_LIGHT)

    # Daftar kolom
    col_start = len(rows) + 4
    ws.cell(row=col_start, column=1, value="All Columns").font = Font(bold=True, color=BLUE_DARK)

    # Build column list from profile
    col_names = []
    col_types = []

    if isinstance(columns, dict):
        # Dict format from profiler: {col_name: {type: ...}, ...}
        for c_name, c_info in columns.items():
            col_names.append(c_name)
            col_types.append(c_info.get("type", "unknown") if isinstance(c_info, dict) else "unknown")
    elif isinstance(columns, list):
        col_names = columns
        # Determine types from numeric/categorical/date lists
        numeric_set = set(numeric_cols) if isinstance(numeric_cols, list) else set()
        date_set = set(date_cols) if isinstance(date_cols, list) else set()
        for c in col_names:
            if c in numeric_set:
                col_types.append("numeric")
            elif c in date_set:
                col_types.append("date")
            else:
                col_types.append("categorical")

    if col_names:
        col_df = pd.DataFrame({
            "Column": col_names,
            "Type":   col_types,
        })
        _write_dataframe(ws, col_df, start_row=col_start + 1, start_col=1)


# ── Sheet: per analisis ───────────────────────────────────────────────────────

def _write_analysis_sheet(
    wb:     Workbook,
    result: AnalysisResult,
    index:  int,
    guardrail=None,
) -> None:
    sheet_name = _safe_sheet_name(result.analysis_name, index)
    ws         = wb.create_sheet(sheet_name)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30

    current_row = 1

    # ── Judul analisis
    current_row = _write_section_header(ws, current_row, result.analysis_name)
    current_row += 1
    
    # ── Guardrails (if available and failed)
    if guardrail and not guardrail.passed:
        ws.cell(row=current_row, column=1, value="⚠️ Data Quality Warning").font = Font(
            bold=True, size=11, color=ORANGE
        )
        current_row += 1
        violations = [v.message for v in guardrail.violations if not v.passed]
        
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        g_cell = ws.cell(row=current_row, column=1)
        g_cell.value = "Beberapa asumsi statistik tidak terpenuhi:\n" + "\n".join([f"- {v}" for v in violations])
        g_cell.alignment = Alignment(wrap_text=True, vertical="top")
        g_cell.font = Font(color=ORANGE)
        
        ws.row_dimensions[current_row].height = 15 * (len(violations) + 1)
        current_row += 2

    # ── Summary table (metrics)
    if result.summary:
        ws.cell(row=current_row, column=1, value="Summary").font = Font(
            bold=True, size=11, color=BLUE_MID
        )
        current_row += 1

        for key, val in result.summary.items():
            ws.cell(row=current_row, column=1, value=str(key)).font  = Font(bold=True)
            ws.cell(row=current_row, column=2, value=str(val))
            if current_row % 2 == 0:
                for col in [1, 2]:
                    ws.cell(row=current_row, column=col).fill = PatternFill(
                        "solid", fgColor=BLUE_LIGHT
                    )
            current_row += 1

        current_row += 1

    # ── Data table (hasil detail)
    data_df = _extract_data_table(result)
    if data_df is not None and not data_df.empty:
        ws.cell(row=current_row, column=1, value="Detail Data").font = Font(
            bold=True, size=11, color=BLUE_MID
        )
        current_row += 1
        current_row  = _write_dataframe(ws, data_df, start_row=current_row, start_col=1)
        current_row += 2

    # ── AI Interpretation
    if result.interpretation:
        ws.cell(row=current_row, column=1, value="AI Interpretation").font = Font(
            bold=True, size=11, color=BLUE_MID
        )
        current_row += 1

        # Merge cells untuk teks panjang
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row,   end_column=6
        )
        interp_cell               = ws.cell(row=current_row, column=1)
        interp_cell.value         = result.interpretation
        interp_cell.alignment     = Alignment(wrap_text=True, vertical="top")
        interp_cell.font          = Font(color=TEXT_DARK, size=10)
        interp_cell.fill          = PatternFill("solid", fgColor=BLUE_LIGHT)

        # Hitung tinggi row berdasarkan panjang teks
        line_count = max(4, len(result.interpretation) // 120 + 1)
        ws.row_dimensions[current_row].height = line_count * 15

    # Set lebar kolom B–F
    for col_letter in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col_letter].width = 20


# ── Helpers ───────────────────────────────────────────────────────────────────

def _get_profile_val(profile, key, default=None):
    """Get a value from profile (works with both dict and dataclass)."""
    if isinstance(profile, dict):
        return profile.get(key, default)
    return getattr(profile, key, default)


def _write_section_header(ws, row: int, title: str) -> int:
    """Tulis header section berwarna biru gelap. Return row berikutnya."""
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row,   end_column=6
    )
    cell           = ws.cell(row=row, column=1, value=title)
    cell.font      = Font(name="Calibri", bold=True, size=13, color=WHITE)
    cell.fill      = PatternFill("solid", fgColor=BLUE_DARK)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 28
    return row + 1


def _write_dataframe(
    ws,
    df:         pd.DataFrame,
    start_row:  int,
    start_col:  int,
    max_rows:   int = 5000,
) -> int:
    """
    Tulis DataFrame ke worksheet.
    Return row setelah data selesai ditulis.
    Header row diberi warna biru.
    Alternating row warna abu.
    """
    # Potong kalau terlalu besar
    if len(df) > max_rows:
        df = df.head(max_rows)

    # Header
    for col_idx, col_name in enumerate(df.columns, start_col):
        cell           = ws.cell(row=start_row, column=col_idx, value=str(col_name))
        cell.font      = Font(bold=True, color=WHITE)
        cell.fill      = PatternFill("solid", fgColor=BLUE_MID)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Auto width berdasarkan nama kolom
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(
            ws.column_dimensions[col_letter].width,
            min(len(str(col_name)) + 4, 40)
        )

    ws.row_dimensions[start_row].height = 20

    # Data rows
    for row_idx, row_data in enumerate(df.itertuples(index=False), start_row + 1):
        for col_idx, value in enumerate(row_data, start_col):
            cell        = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font   = Font(color=TEXT_DARK, size=10)
            if (row_idx - start_row) % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=GRAY_LIGHT)

    return start_row + len(df) + 1


def _extract_data_table(result: AnalysisResult) -> Optional[pd.DataFrame]:
    """
    Ekstrak DataFrame dari result.data untuk ditampilkan sebagai tabel detail.
    Tiap tipe analisis punya key yang berbeda.
    """
    data = result.data or {}

    # Coba key umum yang mengandung list of dicts atau DataFrame
    for key in ("grouped_table", "top_items", "forecast_values",
                "top_pairs", "anomalies", "cluster_sizes",
                "coefficients", "top_anomalies", "association_rules"):
        val = data.get(key)
        if val is None:
            continue
        if isinstance(val, pd.DataFrame) and not val.empty:
            return val
        if isinstance(val, list) and val:
            try:
                return pd.DataFrame(val)
            except Exception:
                continue
        if isinstance(val, dict) and val:
            try:
                return pd.DataFrame(
                    list(val.items()), columns=["key", "value"]
                )
            except Exception:
                continue

    return None


def _safe_sheet_name(name: str, index: int) -> str:
    """
    Buat nama sheet yang valid untuk Excel:
    - Maksimal 31 karakter
    - Tidak boleh: [ ] : * ? / \\
    - Tambah prefix nomor supaya unik
    """
    import re
    clean = re.sub(r'[\\/*?:\[\]]', '', name)
    clean = clean.strip()
    prefix = f"{index:02d}_"
    max_len = 31 - len(prefix)
    return prefix + clean[:max_len]


def _apply_thin_border(ws, min_row, min_col, max_row, max_col) -> None:
    thin = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(
        min_row=min_row, max_row=max_row,
        min_col=min_col, max_col=max_col
    ):
        for cell in row:
            cell.border = border
